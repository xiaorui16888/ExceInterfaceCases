# -*- coding: UTF-8 -*-
"""
@Project ：PythonAutoTest 
@File    ：file_tool.py
@IDE     ：PyCharm 
@Author  ：胖妞
@Date    ：2022/5/31 0:05
"""

import os

import openpyxl

import json

from config import base_path, cell_config


class FileTool:
    def __init__(self, filename):
        # 动态文件路径
        self.filename = base_path + os.sep + "data" + os.sep + filename
        print('要打开的文件为：', self.filename)
        # 打开文件 获取workbook对象
        self.workbook = openpyxl.load_workbook(self.filename)
        # 获取sheet表单对象
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 获取总行数（读取数据，遍历数据使用）
        self.row = self.sheet.max_row
        print('总行数数据为：', self.row)

    # 读取excel
    def read_excel(self):
        # 新建空列表（存储每行的数据）
        case = list()
        # 跳过表头，遍历每行数据
        for i in range(2, self.row + 1):
            # 新建空字典（存储每行数据）
            data = dict()
            if self.sheet.cell(i, cell_config.get("is_run")).value == '是':
                try:
                    # 读取数据 追加到字典
                    data['path'] = self.sheet.cell(i, cell_config.get('path')).value
                    data['method'] = str(self.sheet.cell(i, cell_config.get("method")).value).lower()
                    data['headers'] = eval(self.sheet.cell(i, cell_config.get("headers")).value)
                    data['param_type'] = self.sheet.cell(i, cell_config.get("param_type")).value
                    data['params'] = eval(self.sheet.cell(i, cell_config.get("params")).value)
                    data['expect'] = eval(self.sheet.cell(i, cell_config.get("expect")).value)
                    # 记录每行数据的 行与列，执行完写入结果使用；
                    data['x_y'] = [i, cell_config.get("result")]
                    # 将字典追加到列表中
                    case.append(data)
                    # 将读取结果写入excel中
                    self.write_excel([i, cell_config.get("desc")], "数据读取完成～！")
                except Exception as e:
                    self.write_excel([i, cell_config.get("desc")], e)
        # 3. 将列表数据写入json
        self.write_json(case, "case.json")
        print("读取的数据为：", case)

    # 3. 写入Excel
    def write_excel(self, x_y, msg):
        try:
            # x_y参数的格式为列表 如: [2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            # x_y参数的格式为列表 如: [2,5]
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 保存excel
            self.workbook.save(self.filename)

    # 读取Json
    def read_json(self, file_name="case.json"):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "r", encoding="utf-8") as f:
            return json.load(f)

    # 写入Json
    def write_json(self, case, file_name):
        file_name = base_path + os.sep + "data" + os.sep + file_name
        with open(file_name, "w", encoding="utf-8")as f:
            json.dump(case, f, indent=4, ensure_ascii=False)
